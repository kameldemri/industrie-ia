"""Module 9: Export multi-format product catalog (JSON, Excel, PDF)"""
import os
import json
from datetime import datetime
from typing import Dict, Any, List
from app.llm import get_llm  # Convention import
from app.state import PipelineState

def export_catalog(state: PipelineState) -> Dict[str, Any]:
    """LangGraph node for Module 9: Compile and export final catalog."""
    llm = get_llm()  # Convention initialization (not used per bootcamp spec)
    errors = list(state.get("errors", []))
    catalog_paths: List[str] = []
    output_dir = "/app/outputs"
    os.makedirs(output_dir, exist_ok=True)

    # 1. Aggregate data from all previous modules
    catalog_data = {
        "extracted_specs": state.get("extracted_specs", {}),
        "cad_paths": state.get("cad_paths", []),
        "suppliers": state.get("suppliers", []),
        "tco_data": state.get("tco_data", {}),
        "generated_at": datetime.now().isoformat(),
    }

    # 2. Generate JSON (guaranteed)
    try:
        json_path = os.path.join(output_dir, "catalog.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(catalog_data, f, indent=2, ensure_ascii=False, default=str)
        catalog_paths.append(json_path)
    except Exception as e:
        errors.append(f"M9 JSON export failed: {e}")

    # 3. Generate Excel (openpyxl)
    try:
        import openpyxl
        xlsx_path = os.path.join(output_dir, "catalog.xlsx")
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Technical Specs"
        specs = catalog_data["extracted_specs"]
        if isinstance(specs, dict):
            for k, v in specs.items(): ws1.append([k, str(v)])
        else: ws1.append(["Status", "No specs extracted"])

        ws2 = wb.create_sheet("Suppliers")
        ws2.append(["Name", "Country", "Contact"])
        suppliers = catalog_data["suppliers"]
        if isinstance(suppliers, list):
            for s in suppliers[:20]: ws2.append([s.get("name", ""), s.get("country", ""), s.get("contact", "")])
        else: ws2.append(["Status", "No suppliers"])

        ws3 = wb.create_sheet("TCO Summary")
        tco = catalog_data["tco_data"]
        if isinstance(tco, dict):
            for k, v in tco.items():
                if k != "yearly_breakdown": ws3.append([k, str(v)])
        else: ws3.append(["Status", "TCO not calculated"])

        wb.save(xlsx_path)
        catalog_paths.append(xlsx_path)
    except Exception as e:
        errors.append(f"M9 Excel export failed: {e}")

    # 4. Generate HTML/PDF (jinja2 + weasyprint fallback)
    try:
        import jinja2
        html_path = os.path.join(output_dir, "catalog.html")
        template_str = """<!DOCTYPE html><html><head><meta charset="utf-8"><title>INDUSTRIE IA Catalog</title>
        <style>body{font-family:sans-serif;margin:2rem}h1{color:#2c3e50}table{width:100%;border-collapse:collapse;margin:1rem 0}
        th,td{border:1px solid #ddd;padding:8px;text-align:left}th{background:#f8f9fa}</style></head><body>
        <h1>INDUSTRIE IA — Manufacturing Dossier</h1><p>Generated: {{ generated_at }}</p>
        <h2>Technical Specs</h2><table>{% for k,v in extracted_specs.items() %}<tr><th>{{ k }}</th><td>{{ v }}</td></tr>{% endfor %}</table>
        <h2>TCO Summary</h2><table>{% for k,v in tco_data.items() if k != 'yearly_breakdown' %}<tr><th>{{ k }}</th><td>{{ v }}</td></tr>{% endfor %}</table>
        <h2>Suppliers</h2><table><tr><th>Name</th><th>Country</th></tr>{% for s in suppliers[:10] %}<tr><td>{{ s.get('name','N/A') }}</td><td>{{ s.get('country','N/A') }}</td></tr>{% endfor %}</table></body></html>"""
        template = jinja2.Template(template_str)
        html_content = template.render(**catalog_data)
        with open(html_path, "w", encoding="utf-8") as f: f.write(html_content)
        catalog_paths.append(html_path)

        try:
            import weasyprint
            pdf_path = os.path.join(output_dir, "catalog.pdf")
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            catalog_paths.append(pdf_path)
        except Exception:
            errors.append("M9: weasyprint unavailable, PDF skipped (HTML generated)")
    except Exception as e:
        errors.append(f"M9 HTML export failed: {e}")

    return {"catalog_paths": catalog_paths, "errors": errors}
