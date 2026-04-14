"""
Module 7  Business Plan Generator

input  Module 6 (tco_data)
output:
- JSON business plan
- Excel (.xlsx)
- PDF (.pdf)
includes
- SWOT analysis
- 3 year projections
- ROI
- NPV (VAN)
"""

from __future__ import annotations
import json
from pathlib import Path
from typing import Dict, Any

# ENTRY POINT


def generate_business_plan(state: Dict[str, Any]) -> Dict[str, Any]:

    tco = _load_tco(state)
    product = state.get("extracted_specs") or {}
    suppliers = state.get("suppliers") or []

    plan = _build_plan(tco, product, suppliers)

    Path("outputs").mkdir(exist_ok=True)

    json_path = "outputs/business_plan.json"
    Path(json_path).write_text(json.dumps(plan, indent=2, default=str))

    excel_path = _export_excel(plan, "outputs/business_plan.xlsx")
    pdf_path = _export_pdf(plan, "outputs/business_plan.pdf")
    return {
        "business_plan_paths": {
            "json": json_path,
            "excel": excel_path,
            "pdf": pdf_path,
        },
        "business_plan_summary": {
            "npv": plan["financials"]["npv"],
            "roi_3yr": plan["financials"]["roi_3yr"],
            "revenue_y3": plan["projections"][2]["revenue"],
        },
        "plan": plan
    }



# LOAD TCo

def _load_tco(state: dict) -> dict:
    return state.get("tco_data") or {
        "quantity": 200,
        "unit_material_usd": 450,
        "unit_manufacturing_usd": 200,
        "production_cost_usd": 117000,
        "total_tco_usd": 255000,
    }

# BUSINESS ENGINE


def _build_plan(tco: dict, product: dict, suppliers: list) -> dict:
    quantity = tco["quantity"]
    unit_material = tco.get("unit_material_usd") or tco.get("unit_material", 450)
    unit_manufacturing = tco.get("unit_manufacturing_usd") or tco.get("unit_manufacturing", 200)

    base_cost = unit_material + unit_manufacturing
    unit_cost = base_cost * 0.95
    unit_price = round(unit_cost * 2.5, 2)

    # 3 years
    def year(n):
        revenue = round(unit_price * quantity * (1.2 ** (n - 1)), 2)
        cost = round(unit_cost * quantity * (1.15 ** (n - 1)), 2)
        net = round((revenue - cost) * 0.81, 2)

        return {
            "year": f"Year {n}",
            "revenue": revenue,
            "cost": cost,
            "net_income": net
        }

    projections = [year(1), year(2), year(3)]

    invest = tco["production_cost_usd"]

    # NPV ____________________________
    npv = round(
        sum(p["net_income"] / (1.1 ** (i + 1)) for i, p in enumerate(projections)) - invest,
        2
    )

    #ROI_________________________________
    roi = round(
        ((sum(p["net_income"] for p in projections) - invest) / invest) * 100,
        2
    )

    #SWOT__________________________
    swot = {
        "strengths": [
            "AI cost optimization",
            "Low production cost",
            "Scalable model"
        ],
        "weaknesses": [
            "Estimated pricing only",
            "Limited supplier integration"
        ],
        "opportunities": [
            "Industrial growth",
            "Export potential",
            "Automation demand"
        ],
        "threats": [
            "Competition",
            "Inflation",
            "Supply chain risks"
        ]
    }

    return {
        "product": product.get("product_name", "Unknown"),
        "quantity": quantity,
        "projections": projections,
        "swot": swot,
        "financials": {
            "unit_cost": round(unit_cost, 2),
            "unit_price": unit_price,
            "invest": invest,
            "npv": npv,
            "roi_3yr": roi,
            "total_tco": tco["total_tco_usd"],
        }
    }



# EXCEL EXPORT


def _export_excel(plan: dict, path: str):

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Business Plan"

    ws.append(["Metric", "Value"])
    ws.append(["Product", plan["product"]])
    ws.append(["NPV", plan["financials"]["npv"]])
    ws.append(["ROI", plan["financials"]["roi_3yr"]])
    ws.append(["Total TCO", plan["financials"]["total_tco"]])

    ws2 = wb.create_sheet("Projections")
    ws2.append(["Year", "Revenue", "Cost", "Net Income"])
    for p in plan["projections"]:
        ws2.append([p["year"], p["revenue"], p["cost"], p["net_income"]])

    ws3 = wb.create_sheet("SWOT")
    for k, v in plan["swot"].items():
        ws3.append([k.upper()])
        for item in v:
            ws3.append([item])
        ws3.append([])

    wb.save(path)
    return path



# PDF EXPORT


def _export_pdf(plan: dict, path: str):

    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(path)
    styles = getSampleStyleSheet()

    content = []

    content.append(Paragraph(f"Business Plan: {plan['product']}", styles["Title"]))
    content.append(Spacer(1, 12))

    content.append(Paragraph(f"NPV: {plan['financials']['npv']}", styles["Normal"]))
    content.append(Paragraph(f"ROI: {plan['financials']['roi_3yr']}%", styles["Normal"]))
    content.append(Spacer(1, 12))

    content.append(Paragraph("SWOT ANALYSIS", styles["Heading2"]))
    for k, v in plan["swot"].items():
        content.append(Paragraph(f"{k.upper()}: {', '.join(v)}", styles["Normal"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("3-YEAR PROJECTIONS", styles["Heading2"]))
    for p in plan["projections"]:
        content.append(Paragraph(
            f"{p['year']} — Revenue: {p['revenue']} | Cost: {p['cost']} | Net: {p['net_income']}",
            styles["Normal"]
        ))

    doc.build(content)
    return path



# LOCAL TEST


if __name__ == "__main__":

    import os
    import openpyxl

    os.makedirs("outputs", exist_ok=True)

    print("\n RUNNING MODULE 7 MANUALLY !!!\n")

    
    # Load Module 6 output
    
    def load_tco_from_excel(path="outputs/tco_result.xlsx"):
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb["TCO Summary"]

        data = {}
        for row in sheet.iter_rows(values_only=True):
            if row[0] and row[1]:
                data[row[0]] = row[1]

        return {
            "quantity": data.get("Quantity", 200),
            "unit_material_usd": data.get("Unit Material (USD)", 450),
            "unit_manufacturing_usd": data.get("Unit Manufacturing (USD)", 200),
            "production_cost_usd": data.get("Production Cost Year 0 (USD)", 117000),
            "total_tco_usd": data.get("TOTAL TCO 10 years (USD)", 255000),
        }

  
    # Build state
   
    state = {
        "tco_data": load_tco_from_excel(),
        "extracted_specs": {
            "product_name": "Industrial Product X"
        },
        "suppliers": [
            {"name": "Supplier A"},
            {"name": "Supplier B"}
        ]
    }

    
    # RUN MODULE 7
    
    result = generate_business_plan(state)


    # PRINT RESULTS

    print("\n results M7\n")

    print("Product:", result["plan"]["product"])
    print("NPV:", result["business_plan_summary"]["npv"])
    print("ROI:", result["business_plan_summary"]["roi_3yr"])
    print("Revenue Year 3:", result["business_plan_summary"]["revenue_y3"])

    print("\nFILES GENERATED:")
    print("JSON :", result["business_plan_paths"]["json"])
    print("EXCEL:", result["business_plan_paths"]["excel"])
    print("PDF  :", result["business_plan_paths"]["pdf"])

    print("\nFIN\n")